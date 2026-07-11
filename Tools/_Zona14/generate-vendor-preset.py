#!/usr/bin/env python3
"""Generate a Zona-14 shopPreset YAML from a CSV or Excel sheet.

Expected columns:
    id, price, category

Additional columns are ignored.
"""

import argparse
import csv
import sys
from collections import OrderedDict
from pathlib import Path
from typing import Any

import yaml


def load_ids_from_prototypes(search_dir: Path) -> set[str]:
    """Parse all YAML prototypes under `search_dir` and return the set of `id` values."""
    ids: set[str] = set()
    if not search_dir.exists():
        return ids

    for path in search_dir.rglob("*.yml"):
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = yaml.safe_load(f)
        except Exception:
            continue

        if isinstance(data, list):
            for prototype in data:
                if isinstance(prototype, dict) and "id" in prototype:
                    ids.add(str(prototype["id"]))
        elif isinstance(data, dict) and "id" in data:
            ids.add(str(data["id"]))

    return ids


def read_input_items(input_path: Path) -> list[dict[str, Any]]:
    """Read rows from a CSV or Excel file."""
    rows: list[dict[str, Any]] = []
    suffix = input_path.suffix.lower()

    if suffix == ".csv":
        with open(input_path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                rows.append(row)
    elif suffix in (".xlsx", ".xls"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("openpyxl is required for .xlsx files") from exc

        wb = load_workbook(input_path, read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            raise RuntimeError("Worksheet is empty")

        header = [str(cell).strip() for cell in next(ws.iter_rows(values_only=True))]
        for raw in ws.iter_rows(values_only=True):
            if not any(cell is not None for cell in raw):
                continue
            row = {header[i]: raw[i] for i in range(min(len(header), len(raw)))}
            rows.append(row)
    else:
        raise ValueError(f"Unsupported input extension: {suffix}")

    return rows


def generate_preset(
    items: list[tuple[str, int, str]],
    out_path: Path,
    preset_id: str,
    min_sell_price: int,
    currency_id: str | None = None,
) -> list[str]:
    """Write a `shopPreset` YAML and return missing ID warnings."""
    categories: OrderedDict[str, OrderedDict[str, int]] = OrderedDict()
    missing: list[str] = []

    for item_id, price, category in items:
        if category not in categories:
            categories[category] = OrderedDict()
        # Last price wins if an ID appears twice in the same category.
        categories[category][item_id] = int(price)

    yaml_categories: list[dict[str, Any]] = []
    for priority, (category_name, items_dict) in enumerate(categories.items(), start=1):
        yaml_categories.append(
            {
                "name": category_name,
                "priority": priority,
                "items": dict(items_dict),
            }
        )

    output = {
        "type": "shopPreset",
        "id": preset_id,
        "minSellPrice": min_sell_price,
        "categories": yaml_categories,
    }

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        yaml.safe_dump([output], f, sort_keys=False, default_flow_style=False, allow_unicode=True)

    return missing


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate a shopPreset YAML from a CSV/Excel catalog.")
    parser.add_argument("--input", "-i", required=True, type=Path, help="Input CSV or Excel file")
    parser.add_argument("--output", "-o", required=True, type=Path, help="Output YAML file")
    parser.add_argument("--preset-id", "-p", required=True, help="shopPreset id")
    parser.add_argument("--min-sell-price", "-m", type=int, default=999999999, help="minSellPrice")
    parser.add_argument("--currency", "-c", default="Roubles", help="Currency for validation (unused in output)")
    parser.add_argument("--validate-dir", "-v", type=Path, default=Path("Resources/Prototypes/_Zona14"), help="Directory to validate IDs against")
    args = parser.parse_args()

    rows = read_input_items(args.input)
    items: list[tuple[str, int, str]] = []
    for row in rows:
        item_id = str(row.get("id", "")).strip()
        price = row.get("price", 0)
        category = str(row.get("category", "")).strip()
        if not item_id:
            continue
        try:
            price_int = int(price)
        except (ValueError, TypeError):
            print(f"WARN: non-integer price for '{item_id}': {price}", file=sys.stderr)
            continue
        if not category:
            print(f"WARN: missing category for '{item_id}'", file=sys.stderr)
            continue
        items.append((item_id, price_int, category))

    valid_ids = load_ids_from_prototypes(args.validate_dir)
    missing: list[str] = []
    for item_id, _, _ in items:
        if item_id not in valid_ids:
            missing.append(item_id)

    if missing:
        for item_id in missing:
            print(f"WARN: item id not found in {args.validate_dir}: {item_id}", file=sys.stderr)

    generate_preset(
        items=items,
        out_path=args.output,
        preset_id=args.preset_id,
        min_sell_price=args.min_sell_price,
        currency_id=args.currency,
    )

    print(f"Generated {args.output} with {len(items)} items in {len(set(c for _, _, c in items))} categories.")
    if missing:
        print(f"{len(missing)} missing IDs (see warnings above).")

    return 0 if not missing else 1


if __name__ == "__main__":
    sys.exit(main())
